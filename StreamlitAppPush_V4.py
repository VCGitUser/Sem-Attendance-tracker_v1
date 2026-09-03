import os
import io
import datetime
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Semester Attendance Tracker", layout="wide")

# ─── ACCESS CONTROL ───
def check_password():
    if "password_correct" not in st.session_state:
        st.session_state.password_correct = False
        st.session_state.user_role = "viewer"
    if st.session_state.password_correct:
        return True

    st.subheader("🔒 Secure Portal Access Required")
    user_password = st.text_input("Enter Passcode:", type="password")
    
    if st.button("Verify Key"):
        admin_key = st.secrets["access_password"]
        viewer_key = st.secrets.get("viewer_password", "view123") 
        if user_password == admin_key:
            st.session_state.password_correct = True
            st.session_state.user_role = "admin"
            st.rerun()
        elif user_password == viewer_key:
            st.session_state.password_correct = True
            st.session_state.user_role = "viewer"
            st.rerun()
        else:
            st.error("❌ Invalid Key. Access Denied.")
    return False

if not check_password():
    st.stop()

# Header Layout
st.title("🎈 Student-Supervisor Attendance Portal")
role_badge = "🛠️ ADMIN ACCESS" if st.session_state.user_role == "admin" else "👁️ READ-ONLY ACCESS"
st.markdown(f"Collaborative Web Dashboard | **Role:** `{role_badge}`")
st.divider()

TARGET_EXCEL_FILE = "Semester_Attendance_Tracker_v7.xlsx"
LOG_COLUMN = "Attendance Log (P/A/L/OD)"

@st.cache_data
def load_data():
    df = pd.read_excel(TARGET_EXCEL_FILE, sheet_name="Attendance Ledger", skiprows=12)
    df[LOG_COLUMN] = df[LOG_COLUMN].fillna("").astype(str).str.strip()
    df["Comments / Reason for Absence"] = df["Comments / Reason for Absence"].fillna("").astype(str).str.strip()
    return df.dropna(how='all', axis=0)

try:
    base_df = load_data()
    if "working_df" not in st.session_state:
        st.session_state.working_df = base_df.copy()

    # Capture edits dynamically from data editor
    if st.session_state.user_role == "admin" and "attendance_editor" in st.session_state and st.session_state.attendance_editor:
        changes_detected = st.session_state.attendance_editor["edited_rows"]
        for row_idx, changes in changes_detected.items():
            for col_name, new_val in changes.items():
                st.session_state.working_df.at[row_idx, col_name] = str(new_val).strip()

    active_df = st.session_state.working_df

    # Live Global Summary Calculations (OD counts as Attended)
    total_conducted = active_df[LOG_COLUMN].isin(['P', 'A', 'L', 'OD']).sum()
    total_p = (active_df[LOG_COLUMN] == "P").sum()
    total_od = (active_df[LOG_COLUMN] == "OD").sum()
    total_absent = (active_df[LOG_COLUMN] == "A").sum()
    total_late = (active_df[LOG_COLUMN] == "L").sum()
    
    # Total Attended includes standard Present and On Duty allocations
    total_attended = total_p + total_od
    rate = (total_attended / total_conducted) if total_conducted > 0 else 0.0
    status_text = "ELIGIBLE ✅" if rate >= 0.75 else "SHORTAGE ⚠️"
    
    # Dashboard KPI Cards
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Sessions Conducted", int(total_conducted))
    col2.metric("Total Attended (P + OD)", int(total_attended), help=f"P: {total_p} | OD: {total_od}")
    col3.metric("Absences / Lates", f"{total_absent} A | {total_late} L")
    col4.metric("Attendance Rate", f"{rate:.1%}", delta=status_text, delta_color="normal" if rate >= 0.75 else "inverse")
        
    if total_od > 0:
        st.info(f"ℹ️ **On Duty Notice:** {total_od} sessions logged as `OD` have been credited to the Attended total.")
    if rate < 0.75:
        st.error(f"⚠️ **Attendance Shortage:** Running rate (`{rate:.1%}`) dropped below **75.0%** threshold!")
    else:
        st.success(f"✅ **Status Safe:** Attendance requirement met (`{rate:.1%}`).")

    st.divider()
    
    # Sidebar Aggregations
    st.sidebar.header("🎯 Subject Metrics Tracker")
    st.sidebar.markdown("---")
    subjects = ["UEC3361", "UGE3386", "UEC3301", "UEC3302", "UEC3303", "UMA3362", "LAB"]
    for subj in subjects:
        m = active_df["Course Module"] == subj
        s_p = (m & (active_df[LOG_COLUMN] == 'P')).sum()
        s_od = (m & (active_df[LOG_COLUMN] == 'OD')).sum()
        s_a = (m & (active_df[LOG_COLUMN] == 'A')).sum()
        
        st.sidebar.subheader(f"📘 {subj}")
        sb1, sb2 = st.sidebar.columns(2)
        sb1.write(f"🟢 **P:** {s_p}")
        sb1.write(f"🟣 **OD:** {s_od}")
        sb2.write(f"🔴 **A:** {s_a}")
        st.sidebar.markdown("---")
        
    # Main Interactive Logging Spreadsheet Grid Layout
    st.subheader("🗓️ Academic Calendar Ledger Logs")
    if st.session_state.user_role == "admin":
        st.info("💡 Double-click any cell in the tracking column to change values.")
        disabled_cols = ["Date", "Day", "Period Slot", "Start Time", "End Time", "Course Module"]
    else:
        st.warning("🔒 View-Only mode active. Editing restricted.")
        disabled_cols = active_df.columns.tolist()
    
    st.data_editor(
        active_df,
        disabled=disabled_cols,
        column_config={
            LOG_COLUMN: st.column_config.SelectboxColumn(LOG_COLUMN, options=["P", "A", "L", "OD"], required=False),
            "Comments / Reason for Absence": st.column_config.TextColumn("Comments / Reason for Absence", width="large")
        },
        use_container_width=True,
        num_rows="fixed",
        key="attendance_editor"
    )

    # ─── ADMINISTRATIVE SAVE ENGINE ───
    if st.session_state.user_role == "admin":
        st.divider()
        st.subheader("🛠️ Administrative Controls")
        ac1, ac2 = st.columns(2)
        
        if ac1.button("💾 Save System Changes permanently", use_container_width=True):
            import openpyxl
            wb = openpyxl.load_workbook(TARGET_EXCEL_FILE)
            ws = wb["Attendance Ledger"]
            
            if "Change History" not in wb.sheetnames:
                wsh = wb.create_sheet(title="Change History")
                wsh.append(["Timestamp", "Row", "Module", "Field", "Old", "New"])
            else:
                wsh = wb["Change History"]
            
            t_stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            chg_cnt = 0
            cols_map = {LOG_COLUMN: 7, "Comments / Reason for Absence": 8}
            
            for idx, row in active_df.iterrows():
                r_num = 14 + idx
                for c_name, c_num in cols_map.items():
                    n_val = str(row[c_name]).strip()
                    o_val = str(base_df.at[idx, c_name]).strip()
                    
                    if n_val != o_val:
                        ws.cell(row=r_num, column=c_num, value=row[c_name])
                        v_old = o_val if o_val else "[Blank]"
                        v_new = n_val if n_val else "[Cleared]"
                        wsh.append([t_stamp, f"Row {r_num}", row.get("Course Module", ""), c_name, v_old, v_new])
                        chg_cnt += 1
                        
            if chg_cnt == 0:
                for idx, row in active_df.iterrows():
                    ws.cell(row=14+idx, column=7, value=row[LOG_COLUMN])
                    ws.cell(row=14+idx, column=8, value=row["Comments / Reason for Absence"])
            
            wb.save(TARGET_EXCEL_FILE)
            st.success(f"🎉 Saved! {chg_cnt} structural ledger changes recorded successfully.")
            st.cache_data.clear()
                
        if ac2.button("🔄 Reset Memory (Discard Unsaved Edits)", use_container_width=True):
            if "working_df" in st.session_state:
                del st.session_state.working_df
            st.cache_data.clear()
            st.rerun()
                
        st.markdown("---")
        if os.path.exists(TARGET_EXCEL_FILE):
            with open(TARGET_EXCEL_FILE, "rb") as f:
                st.download_button(
                    label="📥 Download Updated Excel File to Local Computer",
                    data=f.read(),
                    file_name=TARGET_EXCEL_FILE,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
except FileNotFoundError:
    st.error(f"⚠️ Error: Could not locate database file ({TARGET_EXCEL_FILE}).")
